from django.shortcuts import render
from asistencia.models import Asistencia

from datetime import datetime
import time
import openpyxl

import environ
import os

# Create your views here.


def asistencia(request):
    print("Vista: Asistencia")

    id_asistencia = request.user.perfil.id_asistencia

    asistencias = list()
    busqueda_llegada = list()
    busqueda_salida = list()

    if "POST" == request.method:
        fecha = request.POST.get('fecha')
        print("fecha: ", fecha)

        # convertimos la fecha a formato datetime
        fecha = datetime.strptime(fecha, '%Y-%m-%d')
        hora_llegada_minima = datetime(2023, 1, 1, 0, 0, 0).time()
        hora_llegada_maxima = datetime(2023, 1, 1, 11, 0, 0).time()
        hora_salida_minima = datetime(2023, 1, 1, 11, 0, 1).time()
        hora_salida_maxima = datetime(2023, 1, 1, 23, 59, 59).time()

        # unir fecha y hora
        # consultar del modelo Asistencia los registros con el idUser = id_asistencia y la fecha_hora mayor o igual a 2023-01-01 00:00:00
        asistencias = Asistencia.objects.filter(
            idUsuario=id_asistencia, fecha_hora__gte=datetime(2023, 1, 1, 0, 0, 0))

        # busqueda llegada
        busqueda_llegada = Asistencia.objects.filter(
            idUsuario=id_asistencia, fecha_hora__range=[datetime.combine(fecha, hora_llegada_minima), datetime.combine(fecha, hora_llegada_maxima)])
        busqueda_llegada = busqueda_llegada.order_by(
            'fecha_hora').values_list('fecha_hora', flat=True)

        # busqueda salida
        busqueda_salida = Asistencia.objects.filter(idUsuario=id_asistencia, fecha_hora__range=[
            datetime.combine(fecha, hora_salida_minima), datetime.combine(fecha, hora_salida_maxima)])
        busqueda_salida = busqueda_salida.order_by(
            'fecha_hora').values_list('fecha_hora', flat=True)

    # consultar del modelo Asistencia los registros con el idUser = id_asistencia
    asistencias = Asistencia.objects.filter(
        idUsuario=id_asistencia).order_by('-fecha_hora')

    # llegadas_semana = obtener_llegadas(id_asistencia, fecha_lunes_pasado())
    # salidas_semana = obtener_salidas(id_asistencia, fecha_lunes_pasado())

    llegadas_semana = obtener_llegadas(id_asistencia, fecha_lunes())
    salidas_semana = obtener_salidas(id_asistencia, fecha_lunes())

    tolerancia_entrada = int(os.environ.get('TOLERANCIA_ENTRADA'))

    print("")
    print("llegadas_semana: ", llegadas_semana)
    print("")
    print("salidas_semana: ", salidas_semana)
    print("")

    context = {
        'asistencias': asistencias,
        'busqueda_llegada': busqueda_llegada,
        'busqueda_salida': busqueda_salida,
        'llegadas_semana': llegadas_semana,
        'salidas_semana': salidas_semana,
        'tolerancia_entrada': tolerancia_entrada,
        'fecha_lunes_pasado': fecha_lunes()
    }

    return render(request, 'asistencia.html', context)


def carga_asistencia_excel(request):
    print("Vista: Cargar Asistencia Excel")

    n_asistencia = 0
    n_asistencia_existente = 0

    if "GET" == request.method:
        return render(request, 'carga_asistencia_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)

        # print(worksheet)
        print("[+] Excel cargado")

        # getting a particular sheet by name out of many sheets
        # worksheet = wb["Sheet"]

        # obtener la primera hoja del archivo excel
        worksheet = wb[wb.sheetnames[0]]

        # validaciones del archivo excel
        # si el excel tiene mas de 1 hoja
        # if len(wb.sheetnames) > 1:
        #     print("El excel tiene mas de 1 hoja")
        #     return render(request, 'carga_asistencia_excel.html', context={'error': 'Error: El excel tiene mas de 1 hoja'})
        # # o si el tiene mas de 6 columnas
        # elif worksheet.max_column > 6:
        #     print("El excel tiene mas de 6 columnas")
        #     return render(request, 'carga_asistencia_excel.html', context={'error': 'Error: El excel tiene mas de 6 columnas'})
        # # o si el excel no tiene la primera columna con el nombre "ID de Usuario"
        # elif wb[wb.sheetnames[0]].cell(row=1, column=1).value != "ID de Usuario":
        #     print("El excel no tiene la primera columna con el nombre ID de Usuario")
        #     return render(request, 'carga_asistencia_excel.html', context={'error': 'Error: El excel no tiene la primera columna con el nombre ID de Usuario'})
        # # o si el excel no tiene la segunda columna con el nombre "Nombre"
        # elif wb[wb.sheetnames[0]].cell(row=1, column=2).value != "Nombre":
        #     print("El excel no tiene la segunda columna con el nombre Nombre")
        #     return render(request, 'carga_asistencia_excel.html', context={'error': 'Error: El excel no tiene la segunda columna con el nombre Nombre'})
        # # o si el excel no tiene la tercera columna con el nombre "Tiempo"
        # elif wb[wb.sheetnames[0]].cell(row=1, column=3).value != "Tiempo":
        #     print("El excel no tiene la tercera columna con el nombre Tiempo")
        #     return render(request, 'carga_asistencia_excel.html', context={'error': 'Error: El excel no tiene la tercera columna con el nombre Tiempo'})

        excel_data = list()
        asistencia_excel = list()

        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            c = 0
            for cell in row:
                row_data.append(str(cell.value))

            if row_data[0] == "None":
                print("usuario vacio")
            else:
                excel_data.append(row_data)

        # usuario a partir de la segunda fila
        fila = 1
        for i in range(0, len(excel_data) - fila):
            asistencia_excel.append(excel_data[i + fila])

        # si el tamaño de usuarios_excel es mayor a 0, guardamos los usuarios en la base de datos
        if len(asistencia_excel) > 0:
            # imprimir en consola los usuarios del excel
            print(asistencia_excel)

            print("Guardando asistencia en la base de datos")
            n_asistencia, n_asistencia_existente = guardar_asistencia_excel(
                asistencia_excel, request)
        else:
            print("[-] No hay asistencias para guardar en la base de datos")

        context = {
            'excel_data': excel_data,
            'asistencia_excel': asistencia_excel,
            'n_asistencia': n_asistencia,
            'n_asistencia_existente': n_asistencia_existente
        }

        # print(asistencia_excel)

        return render(request, 'carga_asistencia_excel.html', context)


def guardar_asistencia_excel(asistencia_excel, request):
    print("Funcion: Guardar Asistencia Excel")

    n_asistencia = 0
    n_asistencia_existente = 0

    for u in asistencia_excel:

        # idUsuario = u[0].replace(" ", "")
        # nombre = u[1]
        # fecha_hora = u[2]

        idUsuario = u[0].replace(" ", "")
        fecha_hora = u[1]

        print("")
        print("idUsuario: ", idUsuario)
        # print("nombre: ", nombre)
        print("fecha: ", fecha_hora)
        print("")

        # convertir fecha a formato datetime
        fecha_hora = datetime.strptime(fecha_hora, '%Y-%m-%d %H:%M:%S')

        # antes de guardar en la base de datos, verificar si fecha_hora ya existe en la base de datos
        if Asistencia.objects.filter(fecha_hora=fecha_hora).exists():
            # print("[-] La asistencia ya existe en la base de datos")

            # incrementamos el contador de asistencias existentes
            n_asistencia_existente += 1

        else:
            # guardar en la base de datos en el modelo Asistencia los daros del excel
            # Asistencia(idUsuario=idUsuario, nombre=nombre,
            #            fecha_hora=fecha_hora).save()
            Asistencia(idUsuario=idUsuario, fecha_hora=fecha_hora).save()
            # print("[+] Asistencia guardada con exito")

            # incrementamos el contador de asistencias
            n_asistencia += 1

    # imprimir en consola el numero de asistencias guardados
    print("[+] Asistencias guardadas correctamente: ", n_asistencia)
    print("[-] Asistencias existentes: ", n_asistencia_existente)

    return n_asistencia, n_asistencia_existente


# ****************************************************************************************************

# funcion para obtener las llegadas de la semana anterior
def obtener_llegadas(id_asistencia, fecha):
    print("Funcion: obtener_llegadas")

    llegadas_semana = list()
    # ciclo de 6 dias
    for d in range(0, 6):
        consulata = Asistencia.objects.filter(
            idUsuario=id_asistencia, fecha_hora__range=[datetime(fecha.year, fecha.month, fecha.day + d, 0, 0, 0), datetime(fecha.year, fecha.month, fecha.day + d, 11, 0, 0)]).order_by('fecha_hora')

        # si el tamaño de la consulta es 0
        if len(consulata) == 0:
            # agregar un None a la lista
            llegadas_semana.append(None)
        else:
            # agregar a la lista
            llegadas_semana.append(consulata.values(
                'fecha_hora')[0]['fecha_hora'])

    # print(llegadas_semana)

    return llegadas_semana

# funcion para obtener las salidas de la semana anterior


def obtener_salidas(id_asistencia, fecha):
    print("Funcion: obtener_salidas")

    salidas_semana = list()
    # ciclo de 6 dias
    for d in range(0, 6):
        consulata = Asistencia.objects.filter(idUsuario=id_asistencia, fecha_hora__range=[
                                              datetime(fecha.year, fecha.month, fecha.day + d, 11, 0, 1), datetime(fecha.year, fecha.month, fecha.day + d, 23, 59, 59)]).order_by('-fecha_hora')

        # si el tamaño de la consulta es 0
        if len(consulata) == 0:
            # agregar un None a la lista
            salidas_semana.append(None)
        else:
            # agregar a la lista
            salidas_semana.append(consulata.values(
                'fecha_hora')[0]['fecha_hora'])

    # print(salidas_semana)

    return salidas_semana

# realiza un funcion que obtenga la fecha del pasado lunes


def fecha_lunes_pasado():
    print("Funcion: fecha_lunes_pasado")

    import datetime

    today = datetime.date.today()  # fecha actual
    last_monday = today - \
        datetime.timedelta(days=today.weekday() + 7)  # lunes pasado

    # obtener el año, mes y día del lunes pasado
    year = last_monday.year
    month = last_monday.month
    day = last_monday.day

    # crear un objeto datetime con year, month y day
    fecha = datetime.datetime(year, month, day)

    return fecha


# realiza una funcion que obtenga la fecha del lunes de la semana actual

def fecha_lunes():
    print("Funcion: fecha_lunes")

    import datetime

    today = datetime.date.today()  # fecha actual
    last_monday = today - \
        datetime.timedelta(days=today.weekday())  # lunes pasado

    # obtener el año, mes y día del lunes pasado
    year = last_monday.year
    month = last_monday.month
    day = last_monday.day

    # crear un objeto datetime con year, month y day
    fecha = datetime.datetime(year, month, day)

    return fecha
